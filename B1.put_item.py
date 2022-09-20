import boto3
import openpyxl
import pandas   as pd



# path = "vijay.xlsx"
# file = openpyxl.load_workbook(path)
# sheet = file.active

# load excel with its path
wrkbk = openpyxl.load_workbook("vijay.xlsx")

sh = wrkbk.active



dynamodb = boto3.resource('dynamodb',region_name = 'ap-south-1',aws_access_key_id="AKIATZHKUJIM7W4P3EMJ",aws_secret_access_key="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")
table = dynamodb.Table('excel_data')   



# iterate through excel and display data
for row in sh.iter_rows(min_row=1, min_col=1, max_row=1,max_col=5):
    for cell in row:
        print(cell.value)
        
        table.put_item(
        Item= {
        'username': 'vijay',
        'password': 'bhaskar',
        "id": cell.value,
        }
)
  
  

	    

import boto3
import openpyxl


path = "vijay.xlsx"
file = openpyxl.load_workbook(path)
sheet = file.active



dynamodb = boto3.resource('dynamodb',region_name = 'ap-south-1',aws_access_key_id="AKIATZHKUJIM7W4P3EMJ",aws_secret_access_key="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")
table = dynamodb.Table('test')   


for cell in sheet.iter_rows(min_row=2,min_col=3,values_only=True):
    print(cell)
    print(cell[0])
    print(cell[1])
    print(cell[2])
    
    table.put_item(
            Item={
            "id" : '1',
            'row1': cell[0],
            'row2': cell[1],
            'row3': cell[2],
        }
)



print(table.item_count)






# for cell in sheet.iter_rows(min_row=1,max_row=3,max_col=3,min_col=1,values_only=True):
#     print(cell)
#     rw1 = cell[0]
#     rw2 = cell[1]
#     rw3 = cell[2]
       
  

#print(table.item_count)